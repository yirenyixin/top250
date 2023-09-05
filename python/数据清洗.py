import openpyxl
import split
import csv

# 打开 Excel 文件
wb = openpyxl.load_workbook('E:\workspace\\top250\python\\file\TOP250.xlsx')
sheet = wb.active


# 输出 A 列的值
# for row in sheet.iter_rows(min_col=1, max_col=1):
#     for cell in row:
#         print(cell.value)

# 定义要写入的数据
data = [
    ["ID","名称","标签"],
]
i=1
# 指定要创建的 CSV 文件的路径
csv_file_path = "example.csv"

# 使用 "w" 模式打开 CSV 文件，创建文件并写入数据
with open(csv_file_path, "w", newline="",encoding="utf-8") as csv_file:
    csv_writer = csv.writer(csv_file)
    for row in sheet.iter_rows(min_row=2,min_col=1, max_col=1):
        for cell in row:
            data.append([i,cell.value,"电影名称"])
            i+=1
    for row in sheet.iter_rows(min_row=2,min_col=5, max_col=5):
        for cell in row:
            # 使用 \n 拆分字符串并获取拆分后的第一个元素，然后去除多余的空格和换行符
            first_value = cell.value.split('\n')[0].strip()
            # 打印第一个元素
            print(first_value)
            # data.append([i,cell.value,"导演名"])
            # i+=1
    # for row in sheet.iter_rows(min_col=5, max_col=5):
    #     for cell in row:
    #         data.append([i,cell.value,"演员名"])
    #         i+=1
    # for row in sheet.iter_rows(min_col=5, max_col=5):
    #     for cell in row:
    #         data.append([i,cell.value,"类型"])
    #         i+=1
with open(csv_file_path, "w", newline="", encoding="utf-8") as csv_file:
    csv_writer = csv.writer(csv_file)
    for row in data:
        csv_writer.writerow(row)

print(f"CSV 文件 {csv_file_path} 创建成功并写入数据。")