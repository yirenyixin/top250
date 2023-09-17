import openpyxl
import csv
import networkx as nx
import matplotlib.pyplot as plt
all_movie_names=[]=set()
def clean(sheet):
    data = [
        ["片名", "导演", "主演", "类型"],
    ]
    header_row = None
    for row in sheet.iter_rows(values_only=True):
        if "片名" in row and "导演" in row and "主演" in row and "类型" in row:
            header_row = row
            break

    if header_row:
        movie_name_idx = header_row.index("片名")
        director_idx = header_row.index("导演")
        actor_idx = header_row.index("主演")
        genre_idx = header_row.index("类型")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            movie_name = row[movie_name_idx]
            all_movie_names.add(movie_name)
            director = row[director_idx].split(' / ')
            actor = row[actor_idx].split(' / ')
            genre = row[genre_idx].split(' / ')
            data.append([movie_name, director, actor, genre])

    return data

def create_graph(cleaned_data):
    global nodes,edges
    G = nx.Graph()
    # 创建集合用于存储所有节点
    all_nodes = set()
    # 添加节点和边
    for row in cleaned_data[1:]:  # 跳过标题行
        movie_name, directors, actors, genres = row
        # 添加电影节点
        G.add_node(movie_name)
        all_nodes.add(movie_name)
        # 添加导演节点和边
        for director in directors:
            G.add_node(director)
            all_nodes.add(director)
            G.add_edge( movie_name,director, weight=1)
        # 添加演员节点和边
        for actor in actors:
            G.add_node(actor)
            all_nodes.add(actor)
            G.add_edge( movie_name,actor, weight=1)
        # 添加类型节点和边
        for genre in genres:
            G.add_node(genre)
            all_nodes.add(genre)
            G.add_edge( movie_name,genre, weight=1)
    # print(G.nodes)
    # print(G.edges)
    return G


# 自定义排序函数
def sort_by_count(item):
    return item[1]["count"]
def recommend_movies(graph, movie_name, num_recommendations=10):
    if movie_name not in graph:
        return "电影不在数据库中，请选择其他电影。"

    # 使用广度遍历查找与给定电影最相关的电影
    #一个字典，用于存储电影推荐信息。键是电影的名称，值是一个字典，包含两个字段："path" 用于存储电影之间的路径，"count" 用于存储路径的数量。
    movie_recommendations = {}
    #一个集合，用于跟踪已经访问过的电影，以避免重复访问
    visited = set()
    queue = [(movie_name, [])]
    while queue:
        current_movie, path = queue.pop(0)
        visited.add(current_movie)
        for neighbor in graph.neighbors(current_movie):
            if neighbor not in visited   :
                new_path = path + [current_movie]
                queue.append((neighbor, new_path))
                if neighbor != movie_name and neighbor in all_movie_names:
                    if neighbor not in movie_recommendations:
                        movie_recommendations[neighbor] = {"path": new_path, "count": 0}
                    movie_recommendations[neighbor]["count"] += 1

    # 排序推荐电影
    sorted_recommendations = sorted(movie_recommendations.items(), key=sort_by_count, reverse=True)[:num_recommendations]
    # for i in sorted_recommendations:
    #     print(i)
    # 打印推荐电影和推荐理由
    for movie, data in sorted_recommendations:
        path = data["path"]
        count = data["count"]
        print(f"电影：{movie} 推荐理由：共有 {count} 条路径，路径为：{' -> '.join(path)}")
        # print(f"电影：{movie} 推荐理由：2部电影属于{}类型、都是{}导演、都有演员{}参演。")

if __name__ == "__main__":
    # 打开 Excel 文件
    wb = openpyxl.load_workbook('D:\workspace\\top\\top250\python\\file\TOP250.xlsx')
    sheet = wb.active
    # 清洗数据
    cleaned_data = clean(sheet)
    # # 创建图
    graph = create_graph(cleaned_data)
    # movie_name=input("请输入要查询的第一个信息：");
    movie_name="春光乍泄"
    recommend_movies(graph, movie_name)
    # print('开始获取')
    # # 获取所有电影之间的路径
    # all_movie_paths = []
    # movies = [node for node in graph.nodes() if isinstance(node, str)]  # 仅选择电影节点
    # for movie1 in movies:
    #     for movie2 in movies:
    #         if movie1 != movie2 and nx.has_path(graph, movie1, movie2):
    #             path = nx.shortest_path(graph, source=movie1, target=movie2)
    #             print("路径从 {} 到 {}:".format(movie1, movie2))
    #             print(path)
    #             all_movie_paths.append(path)
    # print(all_movie_paths)
    # print('开始保存')
    # # 保存所有路径到 CSV 文件
    # with open('movie_path.csv', 'w', newline='', encoding='utf-8') as csv_file:
    #     csv_writer = csv.writer(csv_file)
    #     for path in all_movie_paths:
    #         csv_writer.writerow(path)
    # print('结束')
    #
    # # 绘制图
    # pos = nx.spring_layout(graph)
    # edge_labels = nx.get_edge_attributes(graph, 'relation')
    # nx.draw(graph, pos, with_labels=True, node_size=10)
    # nx.draw_networkx_edge_labels(graph, pos, edge_labels=edge_labels)
    #
    # # 显示图
    # plt.show()