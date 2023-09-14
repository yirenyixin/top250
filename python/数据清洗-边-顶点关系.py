import openpyxl
import csv

# 指定要创建的 CSV 文件的路径
csv_file_path = "movie_relation.csv"
data = [
    ["名称1", "名称2", "关系"],
]
i=1
def check(name,type,sheet):
    global  i
    for row in sheet.iter_rows(min_row=2):
        if(type.__eq__("电影名称")):
            types=row[7].value.split('/')
            for type in types:
                data.append([name,type.replace(" ",""), "属于"])
            actors = row[6].value.split('/')
            for actor in actors:
                data.append([actor.replace(" ",""),name,"出演"])
        elif(type.__eq__("导演")):
            directors = row[4].value.split('/')
            actors = row[6].value.split('/')
            for director in directors:
                if(name.__eq__(director.replace(" ",""))):
                    data.append([name,row[0].value.replace(" ",""), "导演"])
                    for actor in actors:
                        if not name.__eq__(actor.replace(" ","")):
                            data.append([name,actor.replace(" ",""), "合作"])
        elif(type.__eq__("演员")):
            actors = row[6].value.split('/')
            flag=0
            for actor in actors:
                if(name.__eq__(actor.replace(" ",""))):
                    flag=1
            for actor in actors:
                if not name.__eq__(actor.replace(" ","")) and flag==1:
                    # print(name,actor.replace(" ",""))
                    data.append([name,actor.replace(" ",""), "合作"])



def clean():
    # 读取 CSV 文件
    # with open('E:\workspace\\top250\python\\node.csv', 'r', newline='', encoding='utf-8') as csv_file:
    with open('D:\workspace\\top\\top250\python\\node.csv', 'r', newline='', encoding='utf-8') as csv_file:
        csv_reader = csv.reader(csv_file)
        next(csv_reader)  # 跳过标题行
        for row in csv_reader:
            wb = openpyxl.load_workbook('D:\workspace\\top\\top250\python\\file\TOP250.xlsx')
            # wb = openpyxl.load_workbook('E:/workspace/top250/python/file/TOP250.xlsx')
            sheet = wb.active
            if not row[2].__eq__("类型"):
                check(row[1],row[2],sheet)

    return data
def save(unique_data):
    # 使用 "w" 模式打开 CSV 文件，创建文件并写入数据
    with open(csv_file_path, "w", newline="", encoding="utf-8") as csv_file:
        csv_writer = csv.writer(csv_file)
        for row in unique_data:
            csv_writer.writerow(row)
    # 保存到XLSX文件
    xlsx_file_path = "movie_relation.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    for row in unique_data:
        worksheet.append(row)

    workbook.save(xlsx_file_path)
    print(f"数据也保存到XLSX文件：{xlsx_file_path}")
    print("数据清洗完成。")

if __name__ == "__main__":
    # 清洗
    cleaned_data = clean()
    # 去重
    # unique_data_dict = {}  # 使用字典进行去重操作
    # for row in cleaned_data:
    #     if row[2]=="合作":
    #         # 将每行数据转换成元组并排序，以确保字段顺序无关紧要
    #         row_tuple = tuple(sorted(row[:2]))
    #         # row_tuple = tuple(sorted(row[:3]))
    #         print(row_tuple)
    #         unique_data_dict[row_tuple] = row  # 使用元组作为键来构建字典
    # # 将字典的值（唯一行）转换回列表
    # unique_data = list(unique_data_dict.values())
    # 保存
    save(cleaned_data)
