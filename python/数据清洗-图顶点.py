import openpyxl
import csv

# 指定要创建的 CSV 文件的路径
csv_file_path = "node.csv"

def clean(sheet):
    i = 1
    data = [
        ["ID", "名称", "标签"],
    ]

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            data.append([i, cell.value, "电影名称"])
            i += 1
    directors = []
    for row in sheet.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            # 使用 / 拆分字符串并获取拆分后的第一个元素，然后去除多余的空格和换行符
            first_value = cell.value.split('/')
            for director in first_value:
                directors.append(director.strip())  # 使用 strip() 去除空格和换行符

    unique_directors = list(set(directors))
    for director in unique_directors:
        data.append([i, director.replace(" ", ""), "导演"])
        i += 1
    actors = []
    for row in sheet.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            # 使用 / 拆分字符串并获取拆分后的第一个元素，然后去除多余的空格和换行符
            first_value = cell.value.split('/')
            for actor in first_value:
                actors.append(actor.strip())  # 使用 strip() 去除空格和换行符

    unique_actors = list(set(actors))
    for actor in unique_actors:
        data.append([i, actor.replace(" ", ""), "演员"])
        i += 1

    types = []
    for row in sheet.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            # 使用 / 拆分字符串并获取拆分后的第一个元素，然后去除多余的空格和换行符
            first_value = cell.value.split('/')
            for type in first_value:
                types.append(type.strip())  # 使用 strip() 去除空格和换行符

    unique_types = list(set(types))
    for type in unique_types:
        data.append([i, type.replace(" ", ""), "类型"])
        i += 1

    return data
def save(cleaned_data):
    # 使用 "w" 模式打开 CSV 文件，创建文件并写入数据
    with open(csv_file_path, "w", newline="", encoding="utf-8") as csv_file:
        csv_writer = csv.writer(csv_file)
        for row in cleaned_data:
            csv_writer.writerow(row)

    print("数据清洗完成。")


if __name__ == "__main__":
    # 打开 Excel 文件
    wb = openpyxl.load_workbook('D:\workspace\\top\\top250\python\\file\TOP250.xlsx')
    # wb = openpyxl.load_workbook('E:/workspace/top250/python/file/TOP250.xlsx')
    sheet = wb.active
    # 清洗数据
    cleaned_data = clean(sheet)
    # 保存
    save(cleaned_data)
